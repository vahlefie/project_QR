from datetime import date

from constants import (
    DEFAULT_GUEST_STATUS,
    DEFAULT_SUPER_ADMIN_PASSWORD,
    DEMO_GUEST_EXCEL_FILENAME,
    DEMO_GUEST_EXCEL_PATH,
    GUEST_ADDED_BY_MAX_LENGTH,
    GUEST_STATUS_OPTIONS,
    ROLE_SUPER_ADMIN,
)
from extensions import db
from models import DemoGuest, EventArchive, User
from openpyxl import load_workbook
from sqlalchemy import inspect, text
from werkzeug.security import generate_password_hash

FALLBACK_DEMO_GUEST_COUNT = 1000
FALLBACK_DEMO_SOURCE = "generated_demo_seed"
FALLBACK_DEMO_STAFF_NAMES = ("Staff A", "Staff B", "Staff C", "Staff D", "Staff E")

CLEANUP_GUEST_STATUS_SQL = (
    "UPDATE guests SET status = :default_status " "WHERE status IS NULL OR status NOT IN ('Reguler', 'VIP')"
)

CREATE_USER_WITH_PROVINSI_SQL = """
CREATE TABLE user_new (
    id INTEGER NOT NULL,
    username VARCHAR(15),
    nama VARCHAR(35),
    no_hp INTEGER,
    email VARCHAR(30),
    perusahaan VARCHAR(50),
    alamat VARCHAR(100),
    kota VARCHAR(60),
    provinsi VARCHAR(50),
    aktivasi BOOLEAN,
    paket VARCHAR(10),
    tgl_daftar DATE,
    tgl_expired DATE,
    password VARCHAR(255),
    role VARCHAR(20),
    must_reset_password BOOLEAN,
    active_session_token VARCHAR(128),
    is_blocked BOOLEAN,
    blocked_at DATETIME,
    attendance_token_nonce VARCHAR(64),
    attendance_token_generated_at DATETIME,
    PRIMARY KEY (id),
    UNIQUE (username),
    UNIQUE (no_hp),
    UNIQUE (email)
)
"""

INSERT_USER_WITH_PROVINSI_SQL = """
INSERT INTO user_new (
    id, username, nama, no_hp, email, perusahaan, alamat, kota,
    provinsi, aktivasi, paket, tgl_daftar, tgl_expired, password, role, must_reset_password, active_session_token,
    is_blocked, blocked_at, attendance_token_nonce, attendance_token_generated_at
)
SELECT
    id, username, nama, no_hp, email, perusahaan, alamat, kota,
    NULL, aktivasi, paket, tgl_daftar, tgl_expired, password, role, 0, NULL, 0, NULL, NULL, NULL
FROM "user"
"""


# Fungsi untuk mengambil data super admin default.
def get_default_super_admin_data():
    return {
        "username": "admin",
        "nama": "Super Admin",
        "password": DEFAULT_SUPER_ADMIN_PASSWORD,
        "role": ROLE_SUPER_ADMIN,
        "must_reset_password": False,
    }


# Fungsi untuk memastikan schema user_id tamu.
def ensure_guests_user_schema():
    inspector = inspect(db.engine)
    if "guests" not in inspector.get_table_names():
        return

    guest_columns = {column["name"] for column in inspector.get_columns("guests")}
    if "user_id" not in guest_columns:
        db.session.execute(text("ALTER TABLE guests ADD COLUMN user_id INTEGER"))
        db.session.commit()
        guest_columns.add("user_id")

    if "status" not in guest_columns:
        db.session.execute(text("ALTER TABLE guests ADD COLUMN status VARCHAR(10) DEFAULT 'Reguler'"))
        db.session.commit()
        guest_columns.add("status")

    if "kehadiran" not in guest_columns:
        db.session.execute(text("ALTER TABLE guests ADD COLUMN kehadiran DATETIME"))
        db.session.commit()
        guest_columns.add("kehadiran")

    if "added_by" not in guest_columns:
        db.session.execute(text(f"ALTER TABLE guests ADD COLUMN added_by VARCHAR({GUEST_ADDED_BY_MAX_LENGTH})"))
        db.session.commit()
        guest_columns.add("added_by")

    if "edited_by" not in guest_columns:
        db.session.execute(text(f"ALTER TABLE guests ADD COLUMN edited_by VARCHAR({GUEST_ADDED_BY_MAX_LENGTH})"))
        db.session.commit()
        guest_columns.add("edited_by")

    if "verified_by_staff_id" not in guest_columns:
        db.session.execute(text("ALTER TABLE guests ADD COLUMN verified_by_staff_id INTEGER"))
        db.session.commit()
        guest_columns.add("verified_by_staff_id")

    if "verified_by_staff_name" not in guest_columns:
        db.session.execute(text("ALTER TABLE guests ADD COLUMN verified_by_staff_name VARCHAR(35)"))
        db.session.commit()
        guest_columns.add("verified_by_staff_name")

    if "jumlah_orang" not in guest_columns:
        db.session.execute(text("ALTER TABLE guests ADD COLUMN jumlah_orang INTEGER NOT NULL DEFAULT 1"))
        db.session.commit()
        guest_columns.add("jumlah_orang")

    db.session.execute(
        text("UPDATE guests SET jumlah_orang = 1 " "WHERE jumlah_orang IS NULL OR jumlah_orang < 1 OR jumlah_orang > 9")
    )
    db.session.commit()

    db.session.execute(
        text(CLEANUP_GUEST_STATUS_SQL),
        {"default_status": DEFAULT_GUEST_STATUS},
    )
    db.session.commit()

    inspector = inspect(db.engine)
    guest_indexes = {index["name"] for index in inspector.get_indexes("guests")}
    if "ix_guests_user_id" not in guest_indexes:
        db.session.execute(text("CREATE INDEX ix_guests_user_id ON guests (user_id)"))
        db.session.commit()


# Fungsi untuk memastikan schema provinsi user.
def ensure_user_provinsi_schema():
    inspector = inspect(db.engine)
    if "user" not in inspector.get_table_names():
        return

    user_columns = [column["name"] for column in inspector.get_columns("user")]
    if "provinsi" in user_columns:
        return

    db.session.execute(text("PRAGMA foreign_keys=OFF"))
    db.session.execute(text(CREATE_USER_WITH_PROVINSI_SQL))
    db.session.execute(text(INSERT_USER_WITH_PROVINSI_SQL))
    db.session.execute(text('DROP TABLE "user"'))
    db.session.execute(text('ALTER TABLE user_new RENAME TO "user"'))
    db.session.execute(text("PRAGMA foreign_keys=ON"))
    db.session.commit()


# Fungsi untuk memastikan schema reset password user.
def ensure_user_password_reset_schema():
    inspector = inspect(db.engine)
    if "user" not in inspector.get_table_names():
        return

    user_columns = {column["name"] for column in inspector.get_columns("user")}
    if "must_reset_password" not in user_columns:
        db.session.execute(text('ALTER TABLE "user" ADD COLUMN must_reset_password BOOLEAN DEFAULT 0'))
        db.session.commit()


# Fungsi untuk memastikan schema sesi aktif user.
def ensure_user_active_session_schema():
    inspector = inspect(db.engine)
    if "user" not in inspector.get_table_names():
        return

    user_columns = {column["name"] for column in inspector.get_columns("user")}
    if "active_session_token" not in user_columns:
        db.session.execute(text('ALTER TABLE "user" ADD COLUMN active_session_token VARCHAR(128)'))
        db.session.commit()


# Fungsi untuk memastikan schema block akun user/admin.
def ensure_user_block_schema():
    inspector = inspect(db.engine)
    if "user" not in inspector.get_table_names():
        return

    user_columns = {column["name"] for column in inspector.get_columns("user")}
    if "is_blocked" not in user_columns:
        db.session.execute(text('ALTER TABLE "user" ADD COLUMN is_blocked BOOLEAN DEFAULT 0'))
        db.session.commit()
        user_columns.add("is_blocked")
    if "blocked_at" not in user_columns:
        db.session.execute(text('ALTER TABLE "user" ADD COLUMN blocked_at DATETIME'))
        db.session.commit()


# Fungsi untuk memastikan schema token kehadiran user.
def ensure_user_attendance_token_schema():
    inspector = inspect(db.engine)
    if "user" not in inspector.get_table_names():
        return

    user_columns = {column["name"] for column in inspector.get_columns("user")}
    if "attendance_token_nonce" not in user_columns:
        db.session.execute(text('ALTER TABLE "user" ADD COLUMN attendance_token_nonce VARCHAR(64)'))
        db.session.commit()
        user_columns.add("attendance_token_nonce")
    if "attendance_token_generated_at" not in user_columns:
        db.session.execute(text('ALTER TABLE "user" ADD COLUMN attendance_token_generated_at DATETIME'))
        db.session.commit()


# Fungsi untuk memastikan staff schema.
def ensure_staff_schema():
    inspector = inspect(db.engine)
    if "staff" not in inspector.get_table_names():
        return

    staff_columns = {column["name"] for column in inspector.get_columns("staff")}
    if "is_blocked" not in staff_columns:
        db.session.execute(text("ALTER TABLE staff ADD COLUMN is_blocked BOOLEAN DEFAULT 0"))
        db.session.commit()
        staff_columns.add("is_blocked")
    if "blocked_at" not in staff_columns:
        db.session.execute(text("ALTER TABLE staff ADD COLUMN blocked_at DATETIME"))
        db.session.commit()
        staff_columns.add("blocked_at")
    if "block_reason" not in staff_columns:
        db.session.execute(text("ALTER TABLE staff ADD COLUMN block_reason VARCHAR(100)"))
        db.session.commit()
        staff_columns.add("block_reason")
    if "attendance_token_nonce" not in staff_columns:
        db.session.execute(text("ALTER TABLE staff ADD COLUMN attendance_token_nonce VARCHAR(64)"))
        db.session.commit()
        staff_columns.add("attendance_token_nonce")
    if "attendance_token_generated_at" not in staff_columns:
        db.session.execute(text("ALTER TABLE staff ADD COLUMN attendance_token_generated_at DATETIME"))
        db.session.commit()
        staff_columns.add("attendance_token_generated_at")


# Fungsi untuk memastikan schema request verifikasi kehadiran.
def ensure_attendance_verification_request_schema():
    inspector = inspect(db.engine)
    if "attendance_verification_request" not in inspector.get_table_names():
        return

    request_columns = {column["name"] for column in inspector.get_columns("attendance_verification_request")}
    if "target_staff_id" not in request_columns:
        db.session.execute(text("ALTER TABLE attendance_verification_request ADD COLUMN target_staff_id INTEGER"))
        db.session.commit()


# Fungsi untuk memastikan schema payment client.
def ensure_billing_payment_schema():
    inspector = inspect(db.engine)
    if "billing_payment" not in inspector.get_table_names():
        return

    payment_columns = {column["name"] for column in inspector.get_columns("billing_payment")}
    if "payment_time" not in payment_columns:
        db.session.execute(text("ALTER TABLE billing_payment ADD COLUMN payment_time TIME"))
        db.session.commit()
        payment_columns.add("payment_time")
    if "origin_bank" not in payment_columns:
        db.session.execute(text("ALTER TABLE billing_payment ADD COLUMN origin_bank VARCHAR(60)"))
        db.session.commit()
        payment_columns.add("origin_bank")
    if "account_number" not in payment_columns:
        db.session.execute(text("ALTER TABLE billing_payment ADD COLUMN account_number VARCHAR(40)"))
        db.session.commit()
        payment_columns.add("account_number")
    if "payment_type" not in payment_columns:
        db.session.execute(text("ALTER TABLE billing_payment ADD COLUMN payment_type VARCHAR(20)"))
        db.session.commit()
        payment_columns.add("payment_type")
    if "event_name" not in payment_columns:
        db.session.execute(text("ALTER TABLE billing_payment ADD COLUMN event_name VARCHAR(120)"))
        db.session.commit()
        payment_columns.add("event_name")
    if "accounting_entry" in payment_columns:
        try:
            db.session.execute(text("ALTER TABLE billing_payment DROP COLUMN accounting_entry"))
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            print(f"Kolom accounting_entry belum bisa dihapus otomatis: {exc}")


# Fungsi untuk memastikan schema arsip event.
def ensure_event_archive_schema():
    inspector = inspect(db.engine)
    if "event_archive" not in inspector.get_table_names():
        EventArchive.__table__.create(db.engine)
        db.session.commit()
        return

    archive_columns = {column["name"] for column in inspector.get_columns("event_archive")}
    column_definitions = {
        "package_name": "VARCHAR(10)",
        "period_start": "DATE",
        "period_end": "DATE",
        "csv_path": "VARCHAR(255)",
        "tar_path": "VARCHAR(255)",
        "guest_count": "INTEGER DEFAULT 0",
        "status": "VARCHAR(20) DEFAULT 'csv_ready'",
        "created_at": "DATETIME",
        "archived_at": "DATETIME",
    }
    for column_name, column_type in column_definitions.items():
        if column_name not in archive_columns:
            db.session.execute(text(f"ALTER TABLE event_archive ADD COLUMN {column_name} {column_type}"))
            db.session.commit()

    inspector = inspect(db.engine)
    archive_indexes = {index["name"] for index in inspector.get_indexes("event_archive")}
    if "ix_event_archive_user_id" not in archive_indexes:
        db.session.execute(text("CREATE INDEX ix_event_archive_user_id ON event_archive (user_id)"))
        db.session.commit()
    if "ix_event_archive_period_end" not in archive_indexes:
        db.session.execute(text("CREATE INDEX ix_event_archive_period_end ON event_archive (period_end)"))
        db.session.commit()


# Fungsi untuk membaca data dummy tamu dari file Excel ke objek DemoGuest.
def load_demo_guest_rows_from_excel(excel_path=DEMO_GUEST_EXCEL_PATH):
    if not excel_path.exists():
        return []

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip().lower() if value is not None else "" for value in next(rows, ())]
        demo_guests = []
        for row in rows:
            record = dict(zip(headers, row))
            demo_guest = DemoGuest()
            demo_guest.no = record.get("no")
            demo_guest.nama = record.get("nama")
            demo_guest.no_hp = str(record.get("no_hp") or "")
            demo_guest.email = record.get("email")
            demo_guest.status = record.get("status") or DEFAULT_GUEST_STATUS
            demo_guest.kehadiran = str(record.get("kehadiran") or "")
            demo_guest.verifikasi = record.get("verifikasi")
            demo_guest.source_file = DEMO_GUEST_EXCEL_FILENAME
            demo_guests.append(demo_guest)
        return demo_guests
    finally:
        workbook.close()


# Fungsi untuk membuat data demo dashboard saat file Excel dummy tidak tersedia.
def build_fallback_demo_guest_rows(total_rows=FALLBACK_DEMO_GUEST_COUNT):
    demo_guests = []
    for number in range(1, total_rows + 1):
        demo_guest = DemoGuest()
        demo_guest.no = number
        demo_guest.nama = f"Tamu Demo {number:04d}"
        demo_guest.no_hp = f"6281200{number:06d}"
        demo_guest.email = f"tamu.demo.{number:04d}@example.test"
        demo_guest.status = GUEST_STATUS_OPTIONS[1] if number % 5 == 0 else DEFAULT_GUEST_STATUS
        if number % 4 == 0:
            demo_guest.kehadiran = ""
            demo_guest.verifikasi = ""
        else:
            slot_index = (number - 1) % 24
            hour = 8 + (slot_index // 2)
            minute = 30 if slot_index % 2 else 0
            demo_guest.kehadiran = f"{hour:02d}:{minute:02d}"
            demo_guest.verifikasi = FALLBACK_DEMO_STAFF_NAMES[number % len(FALLBACK_DEMO_STAFF_NAMES)]
        demo_guest.source_file = FALLBACK_DEMO_SOURCE
        demo_guests.append(demo_guest)
    return demo_guests


# Fungsi untuk memastikan tabel data dummy dashboard tersedia dan terisi.
def ensure_demo_guest_schema():
    inspector = inspect(db.engine)
    if "demo_guests" not in inspector.get_table_names():
        DemoGuest.__table__.create(db.engine)
        db.session.commit()

    if DemoGuest.query.count() > 0:
        return

    demo_guests = load_demo_guest_rows_from_excel()
    if not demo_guests:
        demo_guests = build_fallback_demo_guest_rows()
    if not demo_guests:
        return

    db.session.bulk_save_objects(demo_guests)
    db.session.commit()


# Fungsi untuk memastikan super admin default.
def ensure_default_super_admin():
    default_account = get_default_super_admin_data()
    account = User.query.filter_by(username=default_account["username"]).first()
    if not account:
        account = User()
        account.username = default_account["username"]
        account.nama = default_account["nama"]
        account.password = generate_password_hash(default_account["password"])
        account.role = default_account["role"]
        account.tgl_daftar = date.today()
        account.must_reset_password = default_account["must_reset_password"]
        db.session.add(account)
        db.session.commit()
        print("User admin berhasil ditambahkan sebagai Super Admin")
        return account

    changed = False
    if account.role != default_account["role"]:
        account.role = default_account["role"]
        changed = True
    if not account.nama:
        account.nama = default_account["nama"]
        changed = True
    if account.must_reset_password:
        account.must_reset_password = default_account["must_reset_password"]
        changed = True

    if changed:
        db.session.commit()

    return account


# Fungsi untuk menginisialisasi database.
def initialize_database():
    db.create_all()
    ensure_user_provinsi_schema()
    ensure_user_password_reset_schema()
    ensure_user_active_session_schema()
    ensure_user_block_schema()
    ensure_user_attendance_token_schema()
    ensure_staff_schema()
    ensure_attendance_verification_request_schema()
    ensure_billing_payment_schema()
    ensure_event_archive_schema()
    ensure_guests_user_schema()
    ensure_demo_guest_schema()
    ensure_default_super_admin()
